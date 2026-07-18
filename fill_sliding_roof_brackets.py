#!/usr/bin/env python3
"""Recursively find supplier workbooks and fill a summary .xlsm by label."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import et_xmlfile  # noqa: F401 — openpyxl 的隐式依赖，需显式 import 让 PyInstaller 打包
from openpyxl import load_workbook


SUPPLIER_COLUMNS = range(5, 14)  # E:M
DEFAULT_ROWS = {"date": 15, "material_basis": 16, "material": 17,
                "purchased": 18, "production": 19, "overhead_profit": 20,
                "packaging": 21, "freight": 24, "tooling": 30}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9\u0370-\u03ff\u4e00-\u9fff]+", "", str(value).lower())


def number(value: Any) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0.0
    return float(value)


def contains(cell_value: Any, *parts: str) -> bool:
    text = norm(cell_value)
    return any(norm(p) in text for p in parts)


def find_label(ws, *parts: str, min_column: int = 1):
    for row in ws.iter_rows():
        for cell in row:
            if cell.column >= min_column and contains(cell.value, *parts):
                return cell
    raise LookupError(f"找不到标签: {parts}")


def find_exact_label(ws, *labels: str, min_column: int = 1):
    wanted = {norm(label) for label in labels}
    for row in ws.iter_rows():
        for cell in row:
            if cell.column >= min_column and norm(cell.value) in wanted:
                return cell
    raise LookupError(f"找不到精确标签: {labels}")


def find_rightmost_label(ws, *parts: str, exact: bool = False):
    """同一标签出现两次时，最右侧的一组是欧元报价区域。"""
    wanted = {norm(part) for part in parts}
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            text = norm(cell.value)
            matched = text in wanted if exact else any(part in text for part in wanted)
            if matched:
                matches.append(cell)
    if not matches:
        raise LookupError(f"找不到标签: {parts}")
    return max(matches, key=lambda cell: (cell.column, -cell.row))


def value_by_exact_label(ws, data_ws, labels: Iterable[str]):
    wanted = {norm(label) for label in labels}
    label = find_rightmost_label(ws, *wanted, exact=True)
    return next_numeric_right(data_ws, data_ws.cell(label.row, label.column), ws.max_column)


def value_at_label_and_column(ws, data_ws, labels: Iterable[str], result_column: int):
    label = find_rightmost_label(ws, *labels, exact=True)
    value = data_ws.cell(label.row, result_column).value
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        raise LookupError(f"标签 {label.coordinate} 在结果列 {result_column} 没有数值")
    return value


def next_numeric_right(ws, cell, max_columns: int = 20):
    for col in range(cell.column + 1, min(ws.max_column, cell.column + max_columns) + 1):
        value = ws.cell(cell.row, col).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
        if isinstance(value, datetime):
            return value
    raise LookupError(f"标签 {cell.coordinate} 右侧没有数值")


def next_value_right(ws, cell, max_columns: int = 20):
    for col in range(cell.column + 1, min(ws.max_column, cell.column + max_columns) + 1):
        value = ws.cell(cell.row, col).value
        if value not in (None, ""):
            return value
    raise LookupError(f"标签 {cell.coordinate} 右侧没有值")


def row_value(ws, data_ws, *parts: str, preferred_cols=(41, 40, 23, 22, 27)):
    label = find_label(ws, *parts)
    for col in preferred_cols:
        value = data_ws.cell(label.row, col).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
    return next_numeric_right(data_ws, data_ws.cell(label.row, label.column))


def section_rows(ws, header_parts: Iterable[str], end_parts: Iterable[str], min_column=1):
    start = find_label(ws, *header_parts, min_column=min_column).row
    end = find_label(ws, *end_parts, min_column=min_column).row
    if end <= start:
        raise LookupError(f"区间顺序异常: {header_parts} -> {end_parts}")
    return range(start + 1, end)


def max_material_price(ws, data_ws):
    header = find_rightmost_label(ws, "raw material price")
    total = find_rightmost_label(ws, "Σ 1", exact=True)
    if total.row <= header.row:
        raise LookupError("欧元区 raw material price 与 Σ 1 的行顺序异常")
    rows = range(header.row + 1, total.row)
    values = [number(data_ws.cell(r, header.column).value) for r in rows]
    values = [v for v in values if v != 0]
    return max(values) if values else 0


def extract_values(path: Path) -> dict[str, Any]:
    formulas = load_workbook(path, data_only=False, read_only=False)
    values = load_workbook(path, data_only=True, read_only=False)
    ws = formulas[formulas.sheetnames[0]]
    ds = values[ws.title]

    date_label = find_rightmost_label(ws, "quotation date")
    date_value = next_value_right(ds, ds.cell(date_label.row, date_label.column), 20)
    material_label = find_rightmost_label(ws, "Σ 1", exact=True)
    purchased_label = find_rightmost_label(ws, "Σ 2", exact=True)
    material = next_numeric_right(ds, ds.cell(material_label.row, material_label.column), ws.max_column)
    purchased = next_numeric_right(ds, ds.cell(purchased_label.row, purchased_label.column), ws.max_column)
    material_result_col = find_rightmost_label(ws, "cost of material", exact=True).column
    production_result_col = find_rightmost_label(ws, "production costs", exact=True).column
    production = value_at_label_and_column(
        ws, ds, ("manufaction costs", "manufacturing costs"), production_result_col)
    setup = value_at_label_and_column(ws, ds, ("set up costs",), production_result_col)
    material_profit = value_at_label_and_column(
        ws, ds, ("+ additional profit on material",), material_result_col)
    overhead = value_at_label_and_column(ws, ds, ("+ overhead",), production_result_col)
    process_profit = value_at_label_and_column(
        ws, ds, ("+ profit on manufcturing process", "+ profit on manufacturing process"),
        production_result_col)
    packaging_total = find_rightmost_label(ws, "costs2, incl. packaging")
    freight_total = find_rightmost_label(ws, "FOB costs3")
    packaging_header = find_rightmost_label(ws, "packaging costs", exact=True)
    freight_header = find_rightmost_label(ws, "shipping costs", exact=True)
    packaging = number(ds.cell(packaging_total.row, packaging_header.column).value)
    freight = number(ds.cell(freight_total.row, freight_header.column).value)
    tooling = value_at_label_and_column(ws, ds, ("Tooling cost",), production_result_col)

    return {
        "date": date_value,
        "material_basis": max_material_price(ws, ds),
        "material": material,
        "purchased": purchased,
        "production": production + setup,
        "overhead_profit": material_profit + overhead + process_profit,
        "packaging": packaging,
        "freight": freight,
        "tooling": tooling,
    }


def build_file_index(root: Path, summary: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and path.resolve() != summary.resolve():
            index.setdefault(norm(path.stem), []).append(path)
    return index


def resolve(index: dict[str, list[Path]], requested: str, search_root: Path) -> Path:
    matches = index.get(norm(Path(requested).stem), [])
    if not matches:
        raise FileNotFoundError(f"递归目录中找不到同名文件: {requested}")
    valid = [path for path in matches if zipfile.is_zipfile(path)]
    if not valid:
        raise RuntimeError("找到同名文件，但都不是有效的 Excel 文件: " + ", ".join(map(str, matches)))
    return min(valid, key=lambda p: (len(p.relative_to(search_root).parts), str(p).lower()))


def main() -> int:
    if getattr(sys, 'frozen', False):
        script_dir = Path(sys.argv[0]).resolve().parent
    else:
        script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="按名称从供应商报价表提取数据并回填汇总 xlsm")
    parser.add_argument(
        "summary", type=Path, nargs="?",
        help="汇总工作簿；不传时读取脚本同目录下的 Summary_RFQ comparison.xlsm")
    parser.add_argument("--search-root", type=Path, help="递归搜索目录；默认主工作簿所在目录")
    parser.add_argument("--output", type=Path, help="输出 xlsm；默认在原文件名后加 _filled")
    parser.add_argument("--sheet", action="append", help="只处理指定 sheet，可重复；默认处理全部 sheet")
    parser.add_argument("--log", type=Path, help="JSON 日志路径")
    args = parser.parse_args()

    summary = (args.summary or (script_dir / "Summary_RFQ comparison.xlsm")).resolve()
    if not summary.is_file():
        parser.error(f"找不到汇总工作簿: {summary}\n请把脚本与 Summary_RFQ comparison.xlsm 放在同一目录。")
    search_root = (args.search_root or summary.parent).resolve()
    output = (args.output or summary.with_name(summary.stem + "_filled.xlsm")).resolve()
    if summary == output:
        raise ValueError("输出路径不能覆盖源工作簿")
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(summary, output)

    wb = load_workbook(output, keep_vba=True, data_only=False)
    requested_sheets = args.sheet or wb.sheetnames
    index = build_file_index(search_root, summary)
    log = {"summary": str(summary), "output": str(output), "search_root": str(search_root), "items": []}

    for sheet_name in requested_sheets:
        if sheet_name not in wb.sheetnames:
            log["items"].append({"sheet": sheet_name, "status": "error", "error": "sheet 不存在"})
            continue
        if norm(sheet_name) == "summary":
            continue  # 封面/汇总索引页，不是供应商回填明细页
        ws = wb[sheet_name]
        for col in SUPPLIER_COLUMNS:
            requested = ws.cell(11, col).value
            if not isinstance(requested, str) or not requested.strip():
                continue
            item = {"sheet": sheet_name, "column": ws.cell(11, col).column_letter, "requested": requested}
            try:
                source = resolve(index, requested, search_root)
                extracted = extract_values(source)
                for key, row in DEFAULT_ROWS.items():
                    ws.cell(row, col).value = extracted[key]
                item.update(status="ok", source=str(source), values={k: str(v) if isinstance(v, datetime) else v for k, v in extracted.items()})
            except FileNotFoundError as exc:
                item.update(status="skipped", error=str(exc))
            except Exception as exc:
                item.update(status="error", error=str(exc))
            log["items"].append(item)

    wb.save(output)
    log_path = (args.log or output.with_suffix(".json")).resolve()
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
    ok = sum(i["status"] == "ok" for i in log["items"])
    skipped = sum(i["status"] == "skipped" for i in log["items"])
    errors = sum(i["status"] == "error" for i in log["items"])
    print(f"完成: {ok} 个文件，跳过: {skipped} 个，失败: {errors} 个；输出: {output}；日志: {log_path}")
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
